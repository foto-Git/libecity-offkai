#!/usr/bin/env python3
"""
events.json を読み込んで PDF / PNG / Excel を生成する。
Usage: python3 generate_output.py [events.json]
"""
import sys, json, re
from datetime import datetime
from pathlib import Path
from typing import List

OUTPUT_DIR = Path.home() / "Documents" / "libecity-offkai"


# ── ユーティリティ ────────────────────────────────────────

def parse_sort_key(ev: dict):
    """04/02(木) 07:30〜... → (month, day, hour, minute) でソート"""
    d = ev.get("date", "")
    t = ev.get("time", "")
    m = re.match(r"(\d{2})/(\d{2})", d)
    month, day = (int(m.group(1)), int(m.group(2))) if m else (99, 99)
    tm = re.match(r"(\d{1,2}):(\d{2})", t)
    hour, minute = (int(tm.group(1)), int(tm.group(2))) if tm else (99, 99)
    return (month, day, hour, minute)


def fmt_display(ev: dict) -> str:
    return ev.get("format") or "—"


# ── HTML 生成 ─────────────────────────────────────────────

def build_section_rows(events: List[dict], start_i: int) -> str:
    rows = ""
    for i, ev in enumerate(events, start_i):
        bg = "#f8f9ff" if i % 2 == 0 else "#ffffff"
        fmt = fmt_display(ev)
        fmt_class = {"オンライン": "online", "オフライン": "offline"}.get(fmt, "unknown")
        title = ev.get("title") or "（タイトル不明）"
        url   = ev.get("url") or ""
        title_cell = f'<a href="{url}" class="ev-link">{title}</a>' if url else title
        place     = ev.get("place") or "—"
        date_str  = ev.get("date") or "—"
        time_str  = ev.get("time") or ""
        dt_str    = f"{date_str} {time_str}".strip()
        organizer = ev.get("organizer") or "—"
        status    = ev.get("status") or "—"
        source    = ev.get("source") or "—"
        rows += f"""
        <tr style="background:{bg}">
          <td class="num">{i}</td>
          <td class="title">{title_cell}</td>
          <td class="fmt"><span class="badge {fmt_class}">{fmt}</span></td>
          <td class="loc">{place}</td>
          <td class="dt">{dt_str}</td>
          <td class="org">{organizer}</td>
          <td class="sts">{status}</td>
          <td class="src">{source}</td>
        </tr>"""
    return rows


def build_html(events: List[dict], keywords: List[str], generated_at: datetime) -> str:
    today_str = generated_at.strftime("%Y年%m月%d日 %H:%M")
    kw_str    = "、".join(keywords) if keywords else "—"

    sorted_ev = sorted(events, key=parse_sort_key)

    offline = [e for e in sorted_ev if e.get("format") != "オンライン"]
    online  = [e for e in sorted_ev if e.get("format") == "オンライン"]

    def section_html(label: str, evs: List[dict], start: int, color: str) -> str:
        if not evs:
            return ""
        rows = build_section_rows(evs, start)
        return f"""
<div class="section-header" style="background:{color};">
  {label}（{len(evs)}件）
</div>
<table>
  <thead>
    <tr>
      <th class="num">#</th><th>オフ会名</th><th>形式</th>
      <th>開催場所</th><th>開催日時</th><th>主催者</th>
      <th>受付状況</th><th>情報源</th>
    </tr>
  </thead>
  <tbody>{rows}
  </tbody>
</table>"""

    offline_html = section_html("🏢 オフライン開催", offline, 1, "#1a6b3a")
    online_html  = section_html("💻 オンライン開催", online, len(offline) + 1, "#1a3a6b")

    return f"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{
  font-family: 'Hiragino Kaku Gothic ProN','Hiragino Sans','Yu Gothic UI','Meiryo','Noto Sans CJK JP',sans-serif;
  font-size: 12px; color: #222; background:#fff; padding:24px 28px; min-width:1100px;
}}
.header h1 {{ font-size:20px; font-weight:700; color:#1a1a2e; margin-bottom:8px; }}
.meta {{ display:flex; gap:10px; flex-wrap:wrap; margin-bottom:14px; }}
.meta span {{
  font-size:11px; background:#eef1ff; color:#3a4080;
  border:1px solid #c5cdff; border-radius:20px; padding:3px 12px;
}}
.summary {{
  background:linear-gradient(120deg,#3a4080 0%,#6a5acd 100%);
  color:#fff; padding:10px 18px; border-radius:8px;
  font-size:13px; font-weight:700; margin-bottom:18px;
}}
.section-header {{
  color:#fff; font-weight:700; font-size:13px;
  padding:8px 16px; border-radius:6px 6px 0 0; margin-top:24px;
}}
table {{ width:100%; border-collapse:collapse; box-shadow:0 2px 12px rgba(0,0,0,0.07); margin-bottom:8px; }}
thead tr {{ background:#1a1a2e; color:#fff; }}
thead th {{ padding:9px 8px; text-align:left; font-size:11px; font-weight:700; white-space:nowrap; }}
th.num, td.num {{ text-align:center; width:32px; }}
td.title {{ font-weight:600; color:#1a1a2e; max-width:260px; }}
td.fmt {{ text-align:center; width:80px; }}
td.loc {{ max-width:150px; color:#444; }}
td.dt {{ white-space:nowrap; color:#555; font-size:11px; width:140px; }}
td.org {{ max-width:160px; color:#333; font-size:11px; }}
td.sts {{ font-size:11px; color:#666; width:70px; }}
td.src {{ font-size:10px; color:#999; width:70px; }}
tbody td {{ padding:8px 8px; border-bottom:1px solid #eee; vertical-align:top; line-height:1.55; }}
tbody tr:last-child td {{ border-bottom:none; }}
tbody tr:hover {{ background:#f0f2ff !important; }}
.badge {{ display:inline-block; padding:2px 8px; border-radius:12px; font-size:10px; font-weight:700; white-space:nowrap; }}
.badge.online  {{ background:#dbeafe; color:#1d4ed8; }}
.badge.offline {{ background:#dcfce7; color:#166534; }}
.badge.unknown {{ background:#f3f4f6; color:#6b7280; }}
a.ev-link {{ color:#1a1a2e; text-decoration:underline; word-break:break-all; }}
a.ev-link:hover {{ color:#3a4080; }}
.footer {{ margin-top:14px; text-align:right; font-size:11px; color:#bbb; }}
</style>
</head>
<body>
<div class="header">
  <h1>🏙️ リベシティ　オフ会一覧（Claudeキーワード）</h1>
  <div class="meta">
    <span>生成日時：{today_str}</span>
    <span>検索キーワード：{kw_str}</span>
    <span>対象期間：2026/04/02 〜 2026/06/30</span>
    <span>総件数：{len(events)}件（オフライン {len(offline)}件 / オンライン {len(online)}件）</span>
  </div>
</div>
<div class="summary">
  🔍 検索結果：{len(events)} 件のオフ会が見つかりました
  &nbsp;｜&nbsp; オフライン {len(offline)} 件 &nbsp;｜&nbsp; オンライン {len(online)} 件
</div>
{offline_html}
{online_html}
<div class="footer">Generated by リベシティ オフ会検索システム &nbsp;|&nbsp; {today_str}</div>
</body>
</html>"""


# ── PDF / PNG レンダリング ────────────────────────────────

def render(html: str, stem: str):
    from playwright.sync_api import sync_playwright
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)

        pdf_page = browser.new_page(viewport={"width": 1300, "height": 800})
        pdf_page.set_content(html, wait_until="domcontentloaded")
        pdf_page.wait_for_timeout(600)
        pdf_path = OUTPUT_DIR / (stem + ".pdf")
        pdf_page.pdf(
            path=str(pdf_path), format="A4", landscape=True,
            margin={"top": "10mm", "bottom": "10mm", "left": "8mm", "right": "8mm"},
            print_background=True,
        )
        pdf_page.close()
        print(f"  PDF: {pdf_path}")

        png_page = browser.new_page(viewport={"width": 1400, "height": 900}, device_scale_factor=2)
        png_page.set_content(html, wait_until="domcontentloaded")
        png_page.wait_for_timeout(600)
        png_path = OUTPUT_DIR / (stem + ".png")
        png_page.screenshot(path=str(png_path), full_page=True, type="png")
        png_page.close()
        print(f"  PNG: {png_path}")

        browser.close()


# ── Excel 生成 ──────────────────────────────────────────

def build_excel(events: List[dict], keywords: List[str], generated_at: datetime, stem: str):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.hyperlink import Hyperlink
    except ImportError:
        print("  Excel スキップ（openpyxl が未インストール）")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "オフ会一覧"

    # ── スタイル定義 ──
    header_fill   = PatternFill("solid", fgColor="1A1A2E")
    offline_fill  = PatternFill("solid", fgColor="DCFCE7")  # 薄緑
    online_fill   = PatternFill("solid", fgColor="DBEAFE")  # 薄青
    section_off   = PatternFill("solid", fgColor="1A6B3A")
    section_on    = PatternFill("solid", fgColor="1A3A6B")
    notice_fill   = PatternFill("solid", fgColor="FFF3CD")  # 黄色
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    today_str = generated_at.strftime("%Y年%m月%d日 %H:%M")
    kw_str    = "、".join(keywords) if keywords else "claude"

    # ── 行1: タイトル ──
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = f"🏙️ リベシティ オフ会一覧（キーワード：{kw_str}）　生成：{today_str}　総件数：{len(events)}件"
    c.font = Font(bold=True, size=13, color="1A1A2E")
    c.fill = PatternFill("solid", fgColor="EEF1FF")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── 行2: フィルター案内 ──
    ws.merge_cells("A2:I2")
    c = ws["A2"]
    c.value = (
        "💡 フィルター機能が使えます！　"
        "各列ヘッダー（▼）をクリックすると「形式」「開催場所」「受付状況」などで絞り込めます。　"
        "例：形式＝オフラインだけ表示、場所で都市絞り込み、ステータスで募集中だけ表示"
    )
    c.font = Font(bold=True, size=11, color="7D4E00")
    c.fill = notice_fill
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 36

    # ── 行3: ヘッダー ──
    headers = ["#", "日付", "時間", "オフ会名", "形式", "開催場所", "主催者", "受付状況", "情報源"]
    col_widths = [5, 12, 14, 55, 12, 20, 28, 12, 12]
    for col_i, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=3, column=col_i, value=h)
        c.font = Font(bold=True, size=11, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
        ws.column_dimensions[get_column_letter(col_i)].width = w
    ws.row_dimensions[3].height = 22

    # フィルター設定（ヘッダー行に）
    ws.auto_filter.ref = f"A3:I3"

    # ── データ行 ──
    sorted_ev = sorted(events, key=parse_sort_key)
    offline = [e for e in sorted_ev if e.get("format") != "オンライン"]
    online  = [e for e in sorted_ev if e.get("format") == "オンライン"]

    current_row = 4

    def write_section_header(label, count, fill):
        nonlocal current_row
        ws.merge_cells(f"A{current_row}:I{current_row}")
        c = ws.cell(row=current_row, column=1)
        c.value = f"  {label}（{count}件）"
        c.font = Font(bold=True, size=12, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 22
        current_row += 1

    def write_events(evs, row_fill):
        nonlocal current_row
        for num_i, ev in enumerate(evs, 1):
            url   = ev.get("url") or ""
            title = ev.get("title") or "（タイトル不明）"
            vals  = [
                num_i,
                ev.get("date") or "",
                (ev.get("time") or "").split("〜")[0],
                title,
                fmt_display(ev),
                ev.get("place") or "",
                ev.get("organizer") or "",
                ev.get("status") or "",
                ev.get("source") or "",
            ]
            for col_i, val in enumerate(vals, 1):
                c = ws.cell(row=current_row, column=col_i, value=val)
                c.fill = row_fill
                c.border = border
                c.alignment = Alignment(vertical="top", wrap_text=(col_i == 4))
                if col_i == 1:
                    c.alignment = Alignment(horizontal="center", vertical="top")
                if col_i == 4 and url:
                    c.hyperlink = url
                    c.font = Font(color="0563C1", underline="single", size=10)
                else:
                    c.font = Font(size=10)
            ws.row_dimensions[current_row].height = 18
            current_row += 1

    if offline:
        write_section_header("🏢 オフライン開催", len(offline), section_off)
        write_events(offline, offline_fill)
    if online:
        write_section_header("💻 オンライン開催", len(online), section_on)
        write_events(online, online_fill)

    # ウィンドウ枠の固定（ヘッダー行を固定）
    ws.freeze_panes = "A4"

    xlsx_path = OUTPUT_DIR / (stem + ".xlsx")
    wb.save(str(xlsx_path))
    print(f"  Excel: {xlsx_path}")


# ── つぶやき専用 HTML 生成 ───────────────────────────────

def build_tweet_html(events: List[dict], generated_at: datetime) -> str:
    """つぶやきソースのみの一覧HTML（本文プレビュー付き）"""
    today_str = generated_at.strftime("%Y年%m月%d日 %H:%M")
    sorted_ev = sorted(events, key=parse_sort_key)

    rows = ""
    for i, ev in enumerate(sorted_ev, 1):
        bg = "#f8f9ff" if i % 2 == 0 else "#ffffff"
        fmt = fmt_display(ev)
        fmt_class = {"オンライン": "online", "オフライン": "offline"}.get(fmt, "unknown")
        title = ev.get("title") or "（タイトル不明）"
        url   = ev.get("url") or ""
        title_cell = f'<a href="{url}" class="ev-link">{title}</a>' if url else title
        body  = (ev.get("tweetBody") or "").replace("\n", "<br>")
        date_str  = ev.get("date") or "—"
        time_str  = ev.get("time") or ""
        dt_str    = f"{date_str} {time_str}".strip()
        organizer = ev.get("organizer") or "—"
        rows += f"""
        <tr style="background:{bg}">
          <td class="num">{i}</td>
          <td class="dt">{dt_str}</td>
          <td class="org">{organizer}</td>
          <td class="fmt"><span class="badge {fmt_class}">{fmt}</span></td>
          <td class="title">{title_cell}</td>
          <td class="body">{body}</td>
        </tr>"""

    return f"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{
  font-family: 'Hiragino Kaku Gothic ProN','Hiragino Sans','Yu Gothic UI','Meiryo','Noto Sans CJK JP',sans-serif;
  font-size: 12px; color: #222; background:#fff; padding:24px 28px; min-width:1200px;
}}
.header h1 {{ font-size:20px; font-weight:700; color:#1a1a2e; margin-bottom:8px; }}
.meta {{ display:flex; gap:10px; flex-wrap:wrap; margin-bottom:14px; }}
.meta span {{
  font-size:11px; background:#fff8e1; color:#6d4c00;
  border:1px solid #ffe082; border-radius:20px; padding:3px 12px;
}}
.summary {{
  background:linear-gradient(120deg,#b45309 0%,#d97706 100%);
  color:#fff; padding:10px 18px; border-radius:8px;
  font-size:13px; font-weight:700; margin-bottom:18px;
}}
.section-header {{
  color:#fff; font-weight:700; font-size:13px; background:#b45309;
  padding:8px 16px; border-radius:6px 6px 0 0; margin-top:24px;
}}
table {{ width:100%; border-collapse:collapse; box-shadow:0 2px 12px rgba(0,0,0,0.07); }}
thead tr {{ background:#1a1a2e; color:#fff; }}
thead th {{ padding:9px 8px; text-align:left; font-size:11px; font-weight:700; white-space:nowrap; }}
th.num, td.num {{ text-align:center; width:32px; }}
td.dt {{ white-space:nowrap; color:#555; font-size:11px; width:130px; }}
td.org {{ max-width:140px; color:#333; font-size:11px; }}
td.fmt {{ text-align:center; width:80px; }}
td.title {{ font-weight:600; color:#1a1a2e; max-width:220px; word-break:break-all; }}
td.body {{ font-size:11px; color:#555; line-height:1.6; word-break:break-all; }}
tbody td {{ padding:8px 8px; border-bottom:1px solid #eee; vertical-align:top; }}
tbody tr:last-child td {{ border-bottom:none; }}
tbody tr:hover {{ background:#fffbeb !important; }}
.badge {{ display:inline-block; padding:2px 8px; border-radius:12px; font-size:10px; font-weight:700; white-space:nowrap; }}
.badge.online  {{ background:#dbeafe; color:#1d4ed8; }}
.badge.offline {{ background:#dcfce7; color:#166534; }}
.badge.unknown {{ background:#f3f4f6; color:#6b7280; }}
a.ev-link {{ color:#1a1a2e; text-decoration:underline; word-break:break-all; }}
.footer {{ margin-top:14px; text-align:right; font-size:11px; color:#bbb; }}
</style>
</head>
<body>
<div class="header">
  <h1>🐦 リベシティ　つぶやきから拾ったオフ会・Claude関連投稿</h1>
  <div class="meta">
    <span>生成日時：{today_str}</span>
    <span>検索キーワード：claude</span>
    <span>対象期間：2026/04/02 〜 2026/06/30</span>
    <span>件数：{len(sorted_ev)}件</span>
  </div>
</div>
<div class="summary">
  🐦 つぶやき収集結果：{len(sorted_ev)} 件
  &nbsp;｜&nbsp; カレンダー未掲載のオフ会・Claude活用事例を含みます
</div>
<div class="section-header">🐦 つぶやき一覧（{len(sorted_ev)}件）</div>
<table>
  <thead>
    <tr>
      <th class="num">#</th>
      <th style="width:130px">投稿日時</th>
      <th style="width:140px">投稿者</th>
      <th class="fmt">形式</th>
      <th style="width:220px">タイトル（1行目）</th>
      <th>つぶやき本文</th>
    </tr>
  </thead>
  <tbody>{rows}
  </tbody>
</table>
<div class="footer">Generated by リベシティ オフ会検索システム &nbsp;|&nbsp; {today_str}</div>
</body>
</html>"""


# ── メイン ───────────────────────────────────────────────

def main():
    json_path = sys.argv[1] if len(sys.argv) > 1 else str(OUTPUT_DIR / "events.json")
    with open(json_path, encoding="utf-8") as f:
        raw = json.load(f)

    # events.json は list または {"events": [...]} の両方に対応
    if isinstance(raw, list):
        events = raw
        keywords = ["claude"]
    else:
        events = raw.get("events", [])
        keywords = raw.get("keywords", ["claude"])

    now = datetime.now()
    date_prefix = now.strftime("%y%m%d")
    kw_part = "_".join(keywords) if keywords else "検索結果"
    stem = f"{date_prefix}_{kw_part}_libecityオフ会リスト"

    offline = [e for e in events if e.get("format") != "オンライン"]
    online  = [e for e in events if e.get("format") == "オンライン"]
    print(f"イベント件数: {len(events)} 件（オフライン {len(offline)} / オンライン {len(online)}）")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # ── ① 既存：全件一覧（PDF / PNG / Excel）──
    print("\n📄 全件一覧（カレンダー＋つぶやき）を生成中...")
    html = build_html(events, keywords, now)
    render(html, stem)
    build_excel(events, keywords, now, stem)

    # ── ② つぶやきのみ一覧（PDF / PNG）──
    tweet_events = [e for e in events if e.get("source") == "つぶやき"]
    if tweet_events:
        tweet_stem = f"{date_prefix}_{kw_part}_つぶやきオフ会リスト"
        print(f"\n🐦 つぶやきのみ一覧（{len(tweet_events)}件）を生成中...")
        tweet_html = build_tweet_html(tweet_events, now)
        render(tweet_html, tweet_stem)
        print(f"  （Excelは全件一覧に含まれるため省略）")
    else:
        print("\n🐦 つぶやきイベントが見つかりませんでした。")

    print("\n✅ 完了！")


if __name__ == "__main__":
    main()
